from django.shortcuts import render, get_list_or_404, redirect
from core.models import CATEGORY_CHOICES
from core.signals import assign_director_permissions, assign_manager_permissions, assign_staff_permissions
from .models import File, FileAccessLog, EscalationHistory
from django.http import FileResponse, JsonResponse, HttpResponse
from .forms import FileUploadForm, ProblemCategoryForm, TicketForm
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.db.models import Count, Q, Prefetch
from django.utils.timezone import now
from django.db.models.functions import TruncMonth
from core.models import FileCategory
import os
from collections import Counter
from django.contrib.auth.models import User, Group
from django.utils.decorators import method_decorator
from .forms import  EscalationNoteForm, UserUpdateForm, ProfileUpdateForm, TerminalForm,TerminalUploadForm, VersionControlForm, FileUploadForm, CustomUserCreationForm, LoginForm, OTPForm,TicketEditForm,TicketComment, TicketCommentForm, TicketForm
from django.views import View
import csv
from .models import Customer, Region, Terminal, Unit, SystemUser, Zone, ProblemCategory, VersionControl,VersionComment, Report, Ticket, Profile, EmailOTP,TicketComment
from django.core.mail import send_mail, EmailMultiAlternatives, EmailMessage
from django.utils.html import strip_tags
from django.contrib import messages
from datetime import datetime
from django.utils.dateparse import parse_date
from django.utils import timezone
import calendar
from django.shortcuts import get_object_or_404
from mimetypes import guess_type
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import login as auth_login, authenticate
from django import forms
import random
from django.core.exceptions import PermissionDenied
from core.utils import can_user_access_file
from .utils import is_director  
import json
import pandas as pd
from email.mime.image import MIMEImage
from datetime import datetime, timedelta
from io import BytesIO
from core import models
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.views.decorators.http import require_POST
from core.priority_rules import PRIORITY_MATRIX, determine_priority
from django.conf import settings
from django.urls import reverse
#from .models import ISSUE_CHOICES
from .models import ISSUE_MAPPING

def in_group(user, group_name):
    return user.is_authenticated and (user.is_superuser or user.groups.filter(name=group_name).exists())
    
def is_director(user):
    return in_group(user, 'Director')
    
def is_manager(user):
    return in_group(user, 'Manager')
    
def is_staff(user):
    return in_group(user, 'Staff')

@login_required(login_url='login')    
def admin_dashboard(request):
    query = request.GET.get('q', '').strip()
    users_qs = User.objects.all()
    customers_qs = Customer.objects.all()
    terminals_qs = Terminal.objects.select_related('custodian').all()

    if query:
        users_qs = users_qs.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
        customers_qs = customers_qs.filter(
            Q(name__icontains=query)
        )
        terminals_qs = terminals_qs.filter(
            Q(branch_name__icontains=query)
        )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'assign_overseer_or_custodian':
            customer_id = request.POST.get('customer_id')
            overseer_id = request.POST.get('overseer_id')

            customer = get_object_or_404(Customer, id=customer_id)

            # Assign overseer
            if overseer_id:
                overseer = get_object_or_404(User, id=overseer_id)
                overseer.groups.add(Group.objects.get(name='Customer'))
                customer.overseer = overseer
                customer.save()
                print(f"Assigned overseer: {overseer.username}")
                messages.success(request, f"Overseer updated for {customer.name}.")
            else:
                messages.warning(request, f"No overseer selected for {customer.name}.")

        elif action == 'assign_custodian':
            customer_id = request.POST.get('customer_id')
            terminal_id = request.POST.get('terminal_id')
            custodian_id = request.POST.get('custodian_id')

            customer = get_object_or_404(Customer, id=customer_id)
            terminal = get_object_or_404(Terminal, id=terminal_id)
            custodian = get_object_or_404(User, id=custodian_id)

            # Assign custodian to terminal
            terminal.custodian = custodian
            terminal.save()
            

            # Ensure custodian is part of the right group
            custodian.groups.add(Group.objects.get(name='Customer'))

            # Update profile
            profile, _ = Profile.objects.get_or_create(user=custodian)
            profile.terminal = terminal
            profile.customer = customer
            profile.save()

            terminal.refresh_from_db()
            custodian.refresh_from_db()
            if hasattr(custodian, 'profile'):
                custodian.profile.refresh_from_db()

            messages.success(request, f"Custodian {custodian.username} assigned to {terminal.branch_name}.")

        elif action == 'update_role':
            user_id = request.POST.get('user_id')
            new_role = request.POST.get('new_role')

            if user_id and new_role:
                user = get_object_or_404(User, id=user_id)

                # Check if user is a customer (overseer or custodian)
                is_overseer = Customer.objects.filter(overseer=user).exists()
                is_custodian = Terminal.objects.filter(custodian=user).exists()
                is_customer = is_overseer or is_custodian

                restricted_roles = ['Director', 'Manager', 'Staff']
                if is_customer and new_role in restricted_roles:
                    messages.error(request, f"{user.username} is a Customer (overseer or custodian) and cannot be assigned the role '{new_role}'.")
                elif is_customer and new_role == 'Superuser':
                    messages.error(request, f"{user.username} is a Customer and cannot be made superuser.")
                else:
                    # Remove all in-house roles first
                    inhouse_roles = ['Director', 'Manager', 'Staff']
                    user.groups.remove(*Group.objects.filter(name__in=inhouse_roles))

                    # Assign new role group
                    group = Group.objects.get(name=new_role)
                    user.groups.add(group)

                    # Reassign permissions (manually trigger if needed)
                    if new_role == 'Director':
                        assign_director_permissions(user)
                    elif new_role == 'Manager':
                        assign_manager_permissions(user)
                    elif new_role == 'Staff':
                        assign_staff_permissions(user)

                    user.save()
                    messages.success(request, f"{user.username}'s role updated to {new_role}.")
            else:
                messages.error(request, "User ID or role missing in role update.")

        elif action == 'remove_assignment':
            target_type = request.POST.get('target_type')  # 'overseer', 'custodian', or 'role'
            customer_id = request.POST.get('customer_id')
            terminal_id = request.POST.get('terminal_id', None)
            user_id = request.POST.get('user_id')  # for role removal from user

            if target_type == 'overseer':
                if not customer_id or not customer_id.isdigit():
                    messages.error(request, "Invalid customer ID.")
                else:
                    customer = get_object_or_404(Customer, id=customer_id)
                    customer.overseer = None
                    customer.save()
                    messages.success(request, f"Overseer removed from {customer.name}.")

            elif target_type == 'custodian':
                if not terminal_id or not terminal_id.isdigit():
                    messages.error(request, "Invalid terminal ID.")
                else:
                    terminal = get_object_or_404(Terminal, id=terminal_id)
                    terminal.custodian = None
                    terminal.save()
                    Profile.objects.filter(terminal=terminal).update(terminal=None, customer=None)
                    messages.success(request, f"Custodian removed from terminal {terminal.branch_name}.")

            elif target_type == 'role':
                if not user_id or not user_id.isdigit():
                    messages.error(request, "Invalid user ID for role removal.")
                else:
                    user = get_object_or_404(User, id=user_id)
                    roles_to_remove = ['Director', 'Manager', 'Staff']
                    groups_to_remove = Group.objects.filter(name__in=roles_to_remove)
                    user.groups.remove(*groups_to_remove)
                    user.save()
                    messages.success(request, f"Roles removed from user {user.username}.")



        elif action == 'delete_user':
            user_id = request.POST.get('user_id')
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                    user.delete()
                    messages.success(request, f"User {user.username} deleted successfully.")
                except User.DoesNotExist:
                    messages.error(request, "User not found.")
            else:
                messages.error(request, "No user ID provided for deletion.")


    # ========================
    # Access filtering logic
    # ========================
    tickets_qs = Ticket.objects.none()
    files_qs = File.objects.none()

    if request.user.is_superuser or request.user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
        print("Admin/staff user: showing all data.")
        tickets_qs = Ticket.objects.all()
        files_qs = File.objects.all()
    else:
        profile = getattr(request.user, 'profile', None)

        customer = Customer.objects.filter(overseer=request.user).first()
        if customer:
            print(f"{request.user.username} is Overseer of {customer.name}")
            tickets_qs = Ticket.objects.filter(customer=customer)
            files_qs = File.objects.filter(customer=customer)

        elif profile and profile.terminal and profile.customer:
            customer = Customer.objects.filter(custodian=request.user).first()
            if customer:
                print(f"{request.user.username} is Custodian for {profile.terminal} under {customer.name}")
                tickets_qs = Ticket.objects.filter(customer=customer, terminal=profile.terminal)
                files_qs = File.objects.filter(customer=customer, terminal=profile.terminal)
            else:
                print(f"{request.user.username} is a custodian but not linked to any customer.")
        else:
            print(f"{request.user.username} has no access to dashboard data.")

    # Role categorization
    overseer_ids = Customer.objects.filter(overseer__isnull=False).values_list('overseer', flat=True)
    #custodian_ids = Customer.objects.filter(custodian__isnull=False).values_list('custodian', flat=True)
    custodian_ids = Terminal.objects.filter(custodian__isnull=False).values_list('custodian', flat=True).distinct()


    overseer_users = User.objects.filter(id__in=overseer_ids)
    custodian_users = User.objects.filter(id__in=custodian_ids)
    users_without_roles = User.objects.exclude(id__in=list(overseer_ids) + list(custodian_ids))

    # Prefetch custodian → profile → terminal
    profile_prefetch = Prefetch(
        'custodian__profile',
        queryset=Profile.objects.select_related('terminal')
    )
    
    #terminals_prefetch = Prefetch('terminal_set', queryset=Terminal.objects.all(), to_attr='terminals')
    terminals_prefetch = Prefetch('terminal_set', queryset=Terminal.objects.select_related('custodian'), to_attr='terminals')

    context = {
        'users': users_qs,
        #'customers': Customer.objects.prefetch_related(terminals_prefetch),
        'customers': customers_qs.prefetch_related(terminals_prefetch),
        'terminals': terminals_qs,
        'total_users': User.objects.count(),
        'total_files': files_qs.count(),
        'open_tickets': tickets_qs.filter(status='open').count(),
        'overseers': overseer_users,
        'custodians': custodian_users,
        'users_without_roles': users_without_roles,
    }

    return render(request, 'accounts/admin_dashboard.html', context)

@user_passes_test(is_director)
def manage_file_categories(request):
    categories = FileCategory.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            name = request.POST.get('name')
            icon = request.POST.get('icon')
            if name and icon:  
                FileCategory.objects.create(name=name, icon=icon)
                messages.success(request, f'Category "{name}" created successfully.')
                return redirect('manage_file_categories')

        elif action == 'update':
            category_id = request.POST.get('category_id')
            new_name = request.POST.get('new_name')
            new_icon = request.POST.get('icon')
            category = get_object_or_404(FileCategory, id=category_id)
            
            # Preserve the existing icon if no new icon is selected
            category.name = new_name
            if new_icon:
                category.icon = new_icon  
            category.save()
            messages.success(request, f'Category "{new_name}" updated successfully.')
            return redirect('manage_file_categories')

        elif action == 'delete':
            category_id = request.POST.get('category_id')
            category = get_object_or_404(FileCategory, id=category_id)
            category.delete()
            messages.success(request, f'Category "{category.name}" deleted.')
            return redirect('manage_file_categories')

    return render(request, 'accounts/manage_file_categories.html', {
        'categories': categories
    })


@user_passes_test(is_director)
def create_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name') 
        last_name = request.POST.get('last_name') 
        email = request.POST.get('email')
        password = request.POST.get('password')
        role = request.POST.get('role')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('create_user')

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        user.first_name = first_name
        user.last_name = last_name
        user.save()

        group, _ = Group.objects.get_or_create(name=role)
        user.groups.add(group)

        messages.success(request, f"{role} user created successfully.")
        return redirect('admin_dashboard')

    return render(request, 'accounts/create_user.html')

class RegistrationForm(forms.ModelForm):
    password = forms.CharField(widget=forms.PasswordInput)
    class Meta:
        model = User
        fields = ['username', 'email', 'password']


def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'accounts/register.html', {'form': form})


def login_view(request):
    form = LoginForm()

    if request.method == 'POST':
        form = LoginForm(request.POST)
        
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data["password"]
            user = authenticate(username=username, password=password)
            if user is not None:
                user_roles = list(user.groups.values_list('name', flat=True))
                allowed_roles = ['Director', 'Manager', 'Staff', 'Customer'] 

                if not user.is_superuser and not any(role in allowed_roles for role in user_roles):
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Your account has not been issued a role yet. Kindly be patient for a while as we handle that. Thank you!'
                    })
                request.session['pre_otp_user'] = user.id 
                otp = str(random.randint(100000, 999999))
                EmailOTP.objects.update_or_create(user=user, defaults={'otp': otp, 'created_at': timezone.now()})
                
                # Prepare HTML and plain text content
                subject = 'Your OTP Code'
                html_content = f"""
                    <html>
                    <head>
                        <style>
                        @import url('https://fonts.googleapis.com/css?family=Rubik:400,700&display=swap');
                        body {{
                            font-family: 'Rubik', 'Helvetica Neue', Arial, sans-serif;
                            margin: 0;
                            padding: 0;
                            background: linear-gradient(120deg, #1e3c72 0%, #2a5298 100%);
                            min-height: 100vh;
                        }}
                        .email-container {{
                            max-width: 600px;
                            margin: 60px auto;
                            background: #fff;
                            border-radius: 16px;
                            box-shadow: 0 8px 32px 0 rgba(44,62,80,0.12);
                            overflow: hidden;
                            padding: 0 0 40px 0;
                            animation: fadeIn 1s;
                        }}
                        .header-bar {{
                            width: 100%;
                            height: 52px;
                            background: linear-gradient(90deg,#3498db 30%, #e74c3c 80%);
                            display: flex;
                            align-items: center;
                            justify-content: center;
                        }}
                        .logo {{
                            height: 100%;
                            margin: 16px auto 8px auto;
                            display: block;
                            filter: drop-shadow(0 2px 8px rgba(52,152,219,0.12));
                        }}
                        h2 {{
                            text-align: center;
                            color: #1e3c72;
                            font-size: 30px;
                            font-weight: 700;
                            margin: 16px 0 0 0;
                            letter-spacing: 1px;
                        }}
                        .accent-divider {{
                            width: 56px;
                            height: 4px;
                            background: linear-gradient(90deg, #3498db, #e74c3c);
                            border-radius: 2px;
                            margin: 18px auto 24px auto;
                        }}
                        p {{
                            color: #34495e;
                            font-size: 17px;
                            margin: 20px 0;
                            text-align: center;
                        }}
                        .otp-container {{
                            background: linear-gradient(96deg, #e74c3c 60%, #3498db);
                            margin: 35px auto 25px auto;
                            padding: 30px 20px;
                            border-radius: 12px;
                            max-width: 250px;
                            box-shadow: 0 6px 16px 0 rgba(231,76,60,0.08);
                            text-align: center;
                            border: 0.5px solid #f2f2f2;
                            position: relative;
                            animation: fadeIn 2s;
                        }}
                        .otp-glow {{
                            font-size: 36px;
                            font-weight: bold;
                            letter-spacing: 4px;
                            padding: 14px 34px;
                            background: #fff2f0;
                            color: #e74c3c;
                            border-radius: 10px;
                            margin: 20px 0 12px 0;
                            box-shadow: 0 0 25px 7px rgba(231,76,60,0.09);
                            position: relative;
                            animation: shimmer 2.5s linear infinite;
                        }}
                        @keyframes shimmer {{
                            0% {{ box-shadow: 0 0 20px 4px #fffCC2; }}
                            50% {{ box-shadow: 0 0 38px 7px #ffeabf; }}
                            100% {{ box-shadow: 0 0 20px 4px #fffCC2; }}
                        }}
                        .otp-expiry {{
                            color: #fff;
                            font-size: 15px;
                            margin-top: 18px;
                            font-style: italic;
                        }}
                        .cta-button {{
                            margin: 35px auto 0 auto;
                            display: block;
                            width: max-content;
                            background: linear-gradient(90deg, #3498db, #9b59b6);
                            color: #fff !important;
                            font-size: 20px;
                            text-decoration: none;
                            padding: 18px 38px;
                            border-radius: 8px;
                            font-weight: 700;
                            letter-spacing: 1px;
                            box-shadow: 0 4px 12px 0 rgba(41,128,185,0.14);
                            transition: background 0.25s, transform 0.2s;
                        }}
                        .cta-button:hover {{
                            background: #2a70b8;
                            transform: scale(1.06);
                        }}
                        .footer {{
                            margin-top: 60px;
                            font-size: 14px;
                            color: #98a4b3;
                            text-align: center;
                            padding: 32px 14px 0 14px;
                        }}
                        .footer strong {{ color: #34495e; }}
                        .footer a {{
                            color: #2980b9;
                            text-decoration: none;
                            transition: color 0.25s;
                        }}
                        .footer a:hover {{ color: #e74c3c; }}
                        @keyframes fadeIn {{
                            from {{ opacity: 0;transform: translateY(40px); }}
                            to {{ opacity: 1;transform: translateY(0);  }}
                        }}
                        @media only screen and (max-width: 600px) {{
                            .email-container {{ padding: 0 0 20px 0; border-radius: 0; margin: 0; }}
                            .otp-glow {{ font-size: 26px; padding: 10px 10px; }}
                            .otp-container {{ padding: 15px 5px; }}
                        }}
                        </style>
                    </head>
                    <body>
                        <div class="email-container">
                        <div class="header-bar"></div>
                        <img src="cid:logo" alt="BRITS Logo" class="logo" />
                        <h2>Hi {user.username},</h2>
                        <div class="accent-divider"></div>
                        <p>
                            We received a login request for your account.<br>
                            Please use the One-Time Password (OTP) below to complete your login.
                        </p>
                        <div class="otp-container">
                            <div class="otp-glow">{otp}</div>
                            <div class="otp-expiry">Your code expires in <strong>5 minutes</strong>.</div>
                        </div>
                        <p>
                            For your security, do not share this OTP code with anyone.<br>
                            If you did not make this request, please <a href="https://yourapp.com/security">secure your account</a>.
                        </p>
                        <a href="https://yourapp.com/security" class="cta-button">
                            Review Account Activity
                        </a>
                        <div class="footer">
                            Best regards,<br>
                            <strong>Blue River Technology Solutions</strong><br>
                            <span style="font-size:12px;">✉️ This inbox is not monitored for replies</span>
                        </div>
                        </div>
                    </body>
                    </html>
                """
                text_content = strip_tags(html_content)

                # Create the email object
                email = EmailMultiAlternatives(
                    subject,
                    text_content,
                    'no-reply@yourapp.com',
                    [user.email],
                )
                email.attach_alternative(html_content, "text/html")

                # Attach the logo as an inline image using MIMEImage
                """
                with open('static/icons/logo.png', 'rb') as logo_file:
                    logo_data = logo_file.read()
                    logo = MIMEImage(logo_data, name='logo.png')
                    logo.add_header('Content-ID', '<logo>')  

                    email.attach(logo) """
                print("Generated OTP:", otp)

                email.send()
                return JsonResponse({'status': 'otp_sent'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid username or password'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid form input'})
    return render(request, 'accounts/login.html', {'form': form})


def verify_otp_view(request):
    user_id = request.session.get('pre_otp_user')
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'Session expired. Please login again.'})

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'})

    if request.method == "POST":
        form = OTPForm(request.POST)
        if form.is_valid():
            otp_input = form.cleaned_data["otp"]
            otp_instance = EmailOTP.objects.filter(user=user).first()

            if otp_instance:
                if otp_input != otp_instance.otp:
                    return JsonResponse({'status': 'error', 'message': 'Invalid OTP'})
                elif otp_instance.is_expired():
                    return JsonResponse({'status': 'error', 'message': 'Expired OTP'})
                else:
                    auth_login(request, user)
                    if 'pre_otp_user' in request.session:
                        del request.session['pre_otp_user']
                    otp_instance.delete()  
                    return JsonResponse({'status': 'verified', 'redirect_url': '/pre_dashboards/'})
            else:
                return JsonResponse({'status': 'error', 'message': 'OTP not found'})

        return JsonResponse({'status': 'error', 'message': 'Invalid OTP input'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



@login_required
def edit_file(request, file_id):
    #if not request.user.has_perm('core.change_file'):
        #raise PermissionDenied("You do not have permission to edit this file")
    if file.access_level == 'confidential' and file.uploaded_by != request.user and not request.user.is_superuser:
        raise PermissionDenied("This confidential file can only be edited by its uploader.")
    file = get_object_or_404(File, pk=file_id)

    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES, instance=file)
        if form.is_valid():
            form.save()
            return redirect('view_files')
    else:
        form = FileUploadForm(instance=file)

    return render(request, 'edit_file.html', {'form': form, 'file': file})

@login_required(login_url='login')
def pre_dashboards(request):
    # Check if the user is assigned as either overseer or custodian in any customer
    is_overseer_or_custodian = False

    # Check if the user is an overseer or custodian
    if Customer.objects.filter(overseer=request.user).exists() or Terminal.objects.filter(custodian=request.user).exists():
        is_overseer_or_custodian = True

    print(f"Is overseer or custodian: {is_overseer_or_custodian}") 

    return render(request, 'core/pre_dashboards.html', {
        'is_overseer_or_custodian': is_overseer_or_custodian
    })



#@user_passes_test(is_viewer)
def user_list_view(request):
    users = User.objects.all().order_by('username')
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/file_management/user_list.html', {'page_obj': page_obj})

@user_passes_test(is_staff)
def user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    return render(request, 'core/file_management/user_detail.html', {'user': user})

@user_passes_test(is_staff)
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.username = request.POST['username']
        user.email = request.POST['email']
        user.is_active = 'is_active' in request.POST
        user.save()
        messages.success(request, 'User updated successfully!')
        return redirect('user_list')
    return render(request, 'core/file_management/edit_user.html', {'user': user})

@permission_required('auth.delete_user', raise_exception=True)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, 'User deleted successfully!')
    return redirect('user_list')

@login_required
def file_management_dashboard(request):
    # Get all files that are not deleted
    files = File.objects.filter(is_deleted=False)
    
    # Count file extensions
    ext_counter = Counter()
    for f in files:
        ext = os.path.splitext(f.file.name)[1].lower()  # Get file extension
        ext_counter[ext] += 1

    # File types data with counts
    file_types = [
        {"type": "PDF Documents", "ext": ".pdf", "icon": "pdf", "count": ext_counter.get(".pdf", 0)},
        {"type": "Word Documents", "ext": ".docx", "icon": "docx", "count": ext_counter.get(".docx", 0)},
        {"type": "Images", "ext": ".jpg", "icon": "image", "count": ext_counter.get(".jpg", 0) + ext_counter.get(".png", 0)},
        {"type": "Excel Sheets", "ext": ".xlsx", "icon": "xlsx", "count": ext_counter.get(".xlsx", 0)},
        {"type": "PowerPoint", "ext": ".pptx", "icon": "ppt", "count": ext_counter.get(".pptx", 0) + ext_counter.get(".ppt", 0)},
        {"type": "CSV Files", "ext": ".csv", "icon": "csv", "count": ext_counter.get(".csv", 0)},
        {"type": "Text Files", "ext": ".txt", "icon": "text", "count": ext_counter.get(".txt", 0)},
        {"type": "XML Files", "ext": ".xml", "icon": "xml", "count": ext_counter.get(".xml", 0)},
        {"type": "Others", "ext": "other", "icon": "file", "count": sum(ext_counter.values()) - (
            ext_counter.get(".pdf", 0) +
            ext_counter.get(".docx", 0) +
            ext_counter.get(".jpg", 0) +
            ext_counter.get(".png", 0) +
            ext_counter.get(".xlsx", 0) +
            ext_counter.get(".pptx", 0) +
            ext_counter.get(".ppt", 0) +
            ext_counter.get(".csv", 0) +
            ext_counter.get(".txt", 0) +
            ext_counter.get(".xml", 0)
        )},
    ]

    # Count files by category (excluding deleted ones)
    categories = FileCategory.objects.annotate(
        file_count=Count('file', filter=Q(file__is_deleted=False))
    )    

    user = request.user
    # Get recent files, ordered by upload date
    recent_files = File.objects.filter(is_deleted=False).order_by('-upload_date')[:5]
    
    # Check visibility based on access level
    visible_files = []
    for file in recent_files:
        if file.access_level == 'public' or user.is_superuser:
            file.extension = os.path.splitext(file.file.name)[1] 
            visible_files.append(file)
        elif file.access_level == 'restricted' and file.authorized_users.filter(id=user.id).exists() or user.is_superuser:
            file.extension = os.path.splitext(file.file.name)[1]
            visible_files.append(file)
        elif file.access_level == 'confidential' and (file.uploaded_by == user or user.is_superuser):
            file.extension = os.path.splitext(file.file.name)[1]
            visible_files.append(file)

    # Return the render with visible files
    return render(request, 'core/file_management/dashboard.html', {
        'categories': categories,
        'recent_files': visible_files,  
        'file_types': file_types,
        'user_name': request.user.username  
    })

@login_required
def file_list_view(request, category_name=None):
    user = request.user
    files = File.objects.filter(is_deleted=False)

    visible_files = []

    # Iterate over each file to check access rights
    for file in files:
        if file.can_user_access(user):
            visible_files.append(file)

    # Filter by category if provided
    if category_name:
        visible_files = [file for file in visible_files if file.category.name.lower() == category_name.lower()]

    # Sort files
    sort_option = request.GET.get('sort')
    if sort_option == 'recent':
        visible_files.sort(key=lambda f: f.upload_date, reverse=True)
    else:
        visible_files.sort(key=lambda f: f.title)

    # Paginate files
    paginator = Paginator(visible_files, 10)
    page = request.GET.get('page')
    try:
        paginated_files = paginator.page(page)
    except PageNotAnInteger:
        paginated_files = paginator.page(1)
    except EmptyPage:
        paginated_files = paginator.page(paginator.num_pages)

    # Fetch categories for the filter dropdown
    categories = FileCategory.objects.all()

    return render(request, 'core/file_management/file_list.html', {
        'files': paginated_files,
        'categories': categories,
        'active_category': category_name,
    })


def search(request):
    query = request.GET.get('q', '')
    files = File.objects.filter(title__icontains=query, is_deleted=False)
    categories = FileCategory.objects.filter(name__icontains=query)
    users = User.objects.filter(username__icontains=query)

    context = {
        'query': query,
        'files': files,
        'categories': categories,
        'users': users,
    }
    return render(request, 'core/file_management/search_result.html', context)


@login_required
def preview_file(request, file_id):
    file = get_object_or_404(File, id=file_id, is_deleted=False)

    # Access control
    if file.access_level == 'public':
        pass  
    elif file.access_level == 'restricted':
        if not file.can_user_access(request.user):
            raise PermissionDenied("Access denied.")
    elif file.access_level == 'confidential':
        if request.user != file.uploaded_by and not request.user.is_superuser:
            raise PermissionDenied("Access denied for confidential file.")
    else:
        raise PermissionDenied("Unknown access level.")

    # Log access regardless of type
    FileAccessLog.objects.create(file=file, accessed_by=request.user)

    # Serve file if it's previewable
    mime_type, _ = guess_type(file.file.name)
    if mime_type in ['application/pdf', 'image/jpeg', 'image/png', 'image/gif']:
        return FileResponse(file.file.open('rb'), content_type=mime_type)

    # Unsupported type
    return render(request, 'core/file_management/unsupported_preview.html', {'file': file})

@login_required
def delete_file(request, file_id):
    if not request.user.has_perm('core.delete_file'):
        raise PermissionDenied('You do not have permission to delete this file')
    file = get_object_or_404(File, id=file_id, is_deleted=False)

    if request.method == "POST":
        file.is_deleted = True
        file.save()
        messages.success(request, "File deleted successfully.")
        return redirect('file_list')
    
    return redirect('file_list')
    
@login_required
@permission_required('core.add_file', raise_exception=True)
def upload_file_view(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_instance = form.save(commit=False)
            file_instance.uploaded_by = request.user
            file_instance.save()
            return redirect('file_list')
    else:
        form = FileUploadForm()
    
    return render(request, 'core/file_management/upload_file.html', {'form': form})

#@user_passes_test(is_staff)
def profile_view(request):
    context = {
        'user': request.user,
        'user_form': UserUpdateForm(instance=request.user),
        'profile_form': ProfileUpdateForm(instance=request.user.profile),
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'accounts/profile_content.html', context)

    return render(request, 'accounts/profile.html', context)


@method_decorator(login_required, name='dispatch')
class SettingsView(View):
    def get(self, request):
        user_form = UserUpdateForm(instance=request.user)
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile_form = ProfileUpdateForm(instance=profile)
        return render(request, 'accounts/settings.html', {
            'user_form': user_form,
            'profile_form': profile_form
        })

    def post(self, request):
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('profile_view')
        return render(request, 'accounts/settings.html', {
            'user_form': user_form,
            'profile_form': profile_form
        })


@login_required(login_url='login')
def ticketing_dashboard(request):
    now = timezone.localtime(timezone.now())
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = day_start - timedelta(days=day_start.weekday())
    month_start = day_start.replace(day=1)
    year_start = day_start.replace(month=1, day=1)

    # ======================================
    # Role-Based Filtering
    # ======================================
    ticket_filter = Ticket.objects.none()
    profile = getattr(request.user, 'profile', None)

    if request.user.is_superuser or request.user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
        # Superusers and internal staff see all tickets
        ticket_filter = Ticket.objects.all()
    else:
        # For overseer: filter by customer they oversee
        customer = Customer.objects.filter(overseer=request.user).first()
        if customer:
            ticket_filter = Ticket.objects.filter(customer=customer)
        elif profile and profile.terminal:
            # For custodian: filter by the terminal they are assigned to
            # A custodian should see only the tickets associated with their assigned terminal
            ticket_filter = Ticket.objects.filter(terminal=profile.terminal)
            print(f"{request.user.username} is Custodian for {profile.terminal.branch_name} with customer {profile.customer.name}")
        else:
            print(f"{request.user.username} has no profile or terminal")

    # Time-based ticket counts (restricted)
    time_data = {
        'day': ticket_filter.filter(created_at__gte=day_start).count(),
        'week': ticket_filter.filter(created_at__gte=week_start).count(),
        'month': ticket_filter.filter(created_at__gte=month_start).count(),
        'year': ticket_filter.filter(created_at__gte=year_start).count(),
    }

    # Aggregated counts (status & priority)
    status_counts = ticket_filter.values('status').annotate(count=Count('id'))
    priority_counts = ticket_filter.values('priority').annotate(count=Count('id'))

    # Monthly ticket trends
    monthly_trends = (
        ticket_filter
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Top terminals with most tickets
    terminal_data = (
        ticket_filter
        .values('terminal__branch_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Region trends
    region_data = (
        ticket_filter
        .values('terminal__region__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Categories
    category_data = (
        ticket_filter
        .values('problem_category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Top customers
    customer_data = (
        ticket_filter
        .values('terminal__customer__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Overview + Category-time widgets
    overview_data = [
        {'label': 'Daily', 'count': time_data['day']},
        {'label': 'Weekly', 'count': time_data['week']},
        {'label': 'Monthly', 'count': time_data['month']},
        {'label': 'Yearly', 'count': time_data['year']},
    ]

    category_time_data = [
        {'category': d['problem_category__name'], 'daily_count': d['count']} 
        for d in category_data  
    ]

    kpi_data = [
        ("Daily", "dailyCount", "fa-sun"),
        ("Weekly", "weeklyCount", "fa-calendar-week"),
        ("Monthly", "monthlyCount", "fa-calendar-alt"),
        ("Yearly", "yearlyCount", "fa-calendar"),
    ]

    # User Group Determination
    user_group = None
    if Customer.objects.filter(custodian=request.user).exists():
        user_group = "Custodian"
    elif Customer.objects.filter(overseer=request.user).exists():
        user_group = "Overseer"
    else:
        if request.user.groups.filter(name="Director").exists():
            user_group = "Director"
        elif request.user.groups.filter(name="Manager").exists():
            user_group = "Manager"
        elif request.user.groups.filter(name="Staff").exists():
            user_group = "Staff"
        else:
            user_group = "Customer"  

    allowed_roles = ["Director", "Manager", "Staff", "Superuser"]

    context = {
        'user_group': user_group,
        "allowed_roles": allowed_roles,
        "kpi_data": kpi_data,
        'status_data': json.dumps(list(status_counts)),
        'priority_data': json.dumps(list(priority_counts)),
        'monthly_data': json.dumps([
            {'month': calendar.month_abbr[d['month'].month], 'count': d['count']}
            for d in monthly_trends if d['month']
        ]),
        'terminal_data': json.dumps([
            {'terminal': d['terminal__branch_name'], 'count': d['count']}
            for d in terminal_data
        ]),
        'region_data': json.dumps([
            {'region': d['terminal__region__name'], 'count': d['count']}
            for d in region_data
        ]),
        'time_data': json.dumps(time_data),
        'category_data': json.dumps([
            {'category': d['problem_category__name'], 'count': d['count']}
            for d in category_data
        ]),
        'customer_data': json.dumps([
            {'customer': d['terminal__customer__name'], 'count': d['count']}
            for d in customer_data
        ]),
        'overview_data': json.dumps(overview_data),
        'category_time_data': json.dumps(category_time_data),
    }

    return render(request, 'core/helpdesk/ticketing_dashboard.html', context)

@login_required(login_url='login')
def statistics_view(request):
    today = timezone.now()
    print(f"timezone.now() in view: {today} (aware: {timezone.is_aware(today)})")
    # Initialize tickets query to an empty queryset or all, depending on default access policy
    # We will build this query based on user role
    tickets = Ticket.objects.all() # Start with all, then filter down
    user_group = None # To store the identified role for frontend context
    assigned_customer = None
    assigned_terminal = None
    assigned_region = None
    user = request.user
    user_profile = getattr(user, 'profile', None)
    # --- Role-based Filtering (consistent with your working 'tickets' view) ---
    # 1. Internal roles (Superusers, Director, Manager, Staff) - See ALL tickets by default
    if user.is_superuser or user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
        user_group = "Internal" # A more generic term for admin/staff
        print("User has internal access (superuser/staff) - viewing all tickets.")
        # tickets already initialized to Ticket.objects.all()
    # 2. Overseer role: Filter tickets by customer they oversee
    elif Customer.objects.filter(overseer=user).exists():
        user_group = "Overseer"
        # Get the first customer this user oversees (assuming one-to-one or pick one)
        assigned_customer = Customer.objects.filter(overseer=user).first()
        if assigned_customer:
            print(f"{user.username} is Overseer for {assigned_customer.name}")
            tickets = tickets.filter(customer=assigned_customer)
        else:
            # Should not happen if .exists() was true, but good for safety
            tickets = Ticket.objects.none()
            print(f"{user.username} is Overseer but no customer found (unexpected).")
    # 3. Custodian role: Filter tickets by terminal they are assigned to
    # Use profile.terminal and check if user is custodian of that terminal
    elif user_profile and user_profile.terminal:
        # Verify the user is the actual custodian of the terminal assigned to their profile
        if user_profile.terminal.custodian == user:
            user_group = "Custodian"
            assigned_terminal = user_profile.terminal
            assigned_customer = assigned_terminal.customer
            assigned_region = assigned_terminal.region
            print(f"{user.username} is Custodian for terminal {assigned_terminal.branch_name} under customer {assigned_customer.name}")
            tickets = tickets.filter(terminal=assigned_terminal)
        else:
            # User has a terminal in profile, but is not its assigned custodian
            tickets = Ticket.objects.none()
            print(f"{user.username} has terminal in profile but is not its custodian. Returning no tickets.")
    else:
        # Default for users who don't match any of the above roles (e.g., standard users, or those with misconfigured profiles)
        # They should see no tickets by default, or only tickets they created (if that's a general user role)
        tickets = Ticket.objects.none()
        print(f"User {user.username} does not match any specific access role. Returning no tickets.")
    # --- End Role-based Filtering ---
    print(f"Tickets after initial role-based filtering: {tickets.count()}")
    # Get the selected filter values from the request
    time_period = request.GET.get('time-period', "all_time")
    customer_filter = request.GET.get('customer', 'all')
    terminal_filter = request.GET.get('terminal', 'all')
    region_filter = request.GET.get('region', 'all')
    # Determine the date range for filtering based on time period
    # Ensure all dates are timezone-aware if today is
    if time_period == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'yesterday':
        start_date = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'lastweek':
        # This calculates last full week (Sunday to Saturday) relative to today.
        # If today is e.g. Tuesday, it goes back to Sunday of the week before last.
        # For last 7 days including today: start_date = today - timedelta(days=6)
        start_date = today - timedelta(days=today.weekday() + 7) # Monday is 0, Sunday is 6. so add 7 to get to the previous Monday.
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
        # This will get the 7 day period ending yesterday
        # start_date = (today - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        # end_date = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'lastmonth':
        # Go back to the first day of the current month, then subtract one day to get to the end of last month
        end_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
        # Get the first day of last month
        start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif time_period == 'lastyear':
        start_date = today.replace(year=today.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(year=today.year - 1, month=12, day=31, hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'all_time':
        start_date = None
        end_date = None
    else: # Default if time_period is invalid
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    # Apply additional filters based on selected values (customer_filter, terminal_filter, region_filter)
    # IMPORTANT: Ensure these filters don't override the role-based filters in unintended ways.
    # For example, an Overseer should only see *their* customer's data. If customer_filter is applied
    # it must be *within* the scope of the Overseer's assigned customer.
    if customer_filter != 'all':
        # If user is Overseer, ensure customer_filter is within their assigned customer
        # If user is Custodian, their customer is already implicitly filtered
        if user_group == "Overseer" and assigned_customer and str(assigned_customer.id) != customer_filter:
            # If an Overseer tries to filter by a customer they don't oversee, return no tickets
            tickets = Ticket.objects.none()
        elif user_group == "Custodian" and assigned_customer and str(assigned_customer.id) != customer_filter:
             # If a Custodian tries to filter by a customer they don't oversee, return no tickets
            tickets = Ticket.objects.none()
        else:
            tickets = tickets.filter(terminal__customer__id=customer_filter)
    if terminal_filter != 'all':
        # If user is Custodian, ensure terminal_filter matches their assigned terminal
        if user_group == "Custodian" and assigned_terminal and str(assigned_terminal.id) != terminal_filter:
            tickets = Ticket.objects.none()
        else:
            tickets = tickets.filter(terminal__id=terminal_filter)
    if region_filter != 'all':
        # For Custodians, their region is already implicitly filtered via terminal.
        # For Overseers, if a region filter is applied, it must be within their customer's regions.
        if user_group == "Custodian" and assigned_region and str(assigned_region.id) != region_filter:
            tickets = Ticket.objects.none()
        else:
            tickets = tickets.filter(terminal__region__id=region_filter)
    if time_period != 'all_time':
        tickets = tickets.filter(created_at__range=[start_date, end_date])
    # Ensure query is executed before doing in-memory filtering for distinct years etc.
    tickets_list = list(tickets.iterator()) # Convert to list to avoid re-querying for each aggregation
    # Calculate ticket statuses
    # Use the filtered 'tickets' queryset directly
    ticket_statuses = tickets.values('status').annotate(status_count=Count('status'))
    status_labels = [status['status'] for status in ticket_statuses]
    status_counts = [status['status_count'] for status in ticket_statuses]
    # Data aggregation for different timeframes
    # Ensure these aggregations use the already filtered 'tickets' queryset
    days = [today - timedelta(days=i) for i in range(7)]  # Last 7 days including today
    tickets_per_day = [tickets.filter(created_at__date=day.date()).count() for day in days]
    weekdays = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    # Adjust for Django's week_day (1=Sunday, 7=Saturday)
    tickets_per_weekday = [tickets.filter(created_at__week_day=(i % 7) + 1).count() for i in range(7)] # Ensure correct mapping
    hours = [f"{i}-{i+1}" for i in range(24)]
    tickets_per_hour = [tickets.filter(created_at__hour=i).count() for i in range(24)]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    tickets_per_month = [tickets.filter(created_at__month=i+1).count() for i in range(12)]
    # Collect distinct years from the already filtered tickets
    years = sorted(list(set(ticket.created_at.year for ticket in tickets_list))) # Use tickets_list
    tickets_per_year = [tickets.filter(created_at__year=year).count() for year in years]
    tickets_per_terminal = tickets.values('terminal__branch_name').annotate(ticket_count=Count('id'))
    ticket_categories = tickets.values('problem_category__name').annotate(ticket_count=Count('id')) # Use __name for category name
    # Fetch terminals, customers, and regions for the filter options
    # These should be filtered based on the user's access
    available_customers = []
    available_terminals = []
    available_regions = []
    if user_group == "Internal":
        available_customers = list(Customer.objects.values('id', 'name'))
        available_terminals = list(Terminal.objects.select_related('customer', 'region').values(
            'id', 'branch_name', 'customer__name', 'region__name'
        ))
        available_regions = list(Region.objects.values('id', 'name'))
    elif user_group == "Overseer" and assigned_customer:
        available_customers = [{'id': assigned_customer.id, 'name': assigned_customer.name}]
        available_terminals = list(Terminal.objects.filter(customer=assigned_customer).select_related('customer', 'region').values(
            'id', 'branch_name', 'customer__name', 'region__name'
        ))
        available_regions = list(Region.objects.filter(terminal__customer=assigned_customer).distinct().values('id', 'name'))
    elif user_group == "Custodian" and assigned_terminal:
        available_customers = [{'id': assigned_customer.id, 'name': assigned_customer.name}]
        available_terminals = [{
            'id': assigned_terminal.id,
            'branch_name': assigned_terminal.branch_name,
            'customer__name': assigned_terminal.customer.name if assigned_terminal.customer else 'N/A',
            'region__name': assigned_terminal.region.name if assigned_terminal.region else 'N/A',
        }]
        available_regions = [{'id': assigned_region.id, 'name': assigned_region.name}] if assigned_region else []
    else:
        # For users with no specific access, they see no filter options
        available_customers = []
        available_terminals = []
        available_regions = []
    # Map the fetched data to the required format for the frontend
    terminals_for_frontend = [
        {
            'id': t['id'],
            'branch_name': t['branch_name'],
            'customer_name': t['customer__name'],
            'region_name': t['region__name'],
        }
        for t in available_terminals
    ]
    # Pack data for the frontend
    data = {
        'ticketsPerTerminal': [
            {
                'branch_name': entry['terminal__branch_name'],
                'count': entry['ticket_count']
            }
            for entry in tickets_per_terminal
        ],
        'ticketCategories': {
            'labels': [entry['problem_category__name'] for entry in ticket_categories], 
            'data': [entry['ticket_count'] for entry in ticket_categories]
        },
        'ticketStatuses': {
            'labels': status_labels,
            'data': status_counts,
        },
        'days': [day.strftime('%Y-%m-%d') for day in days],
        'ticketsPerDay': tickets_per_day,
        'weekdays': weekdays,
        'ticketsPerWeekday': tickets_per_weekday,
        'hours': hours,
        'ticketsPerHour': tickets_per_hour,
        'months': months,
        'ticketsPerMonth': tickets_per_month,
        'years': years,
        'ticketsPerYear': tickets_per_year,
        'terminals': terminals_for_frontend, 
        'customers': available_customers,     
        'regions': available_regions,         
    }
    # Check if the request is AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(data, safe=False)
    return render(request, 'core/helpdesk/statistics.html', {
        'customers': available_customers, 
        'terminals': terminals_for_frontend, 
        'regions': available_regions,     
        "time_period": time_period,
        'selected_customer': customer_filter,
        'selected_terminal': terminal_filter,
        'selected_region': region_filter,
        'data_json': json.dumps(data, ensure_ascii=False),
        "user_group": user_group,
        "assigned_customer": assigned_customer, 
        "assigned_branch": assigned_terminal,
        "assigned_region": assigned_region,
    })


@login_required(login_url='login')
def export_report(request):
    # Get the selected filter values from the request
    time_period = request.GET.get('time-period', 'today')
    customer_filter = request.GET.get('customer', 'all')
    terminal_filter = request.GET.get('terminal', 'all')
    region_filter = request.GET.get('region', 'all')

    print(f"Export filters: Time period={time_period}, Customer={customer_filter}, Terminal={terminal_filter}, Region={region_filter}")

    # Get the current time (timezone-aware)
    today = timezone.now()

    # Determine the date range for filtering based on time period
    start_date = None
    end_date = None

    if time_period == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'yesterday':
        start_date = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'lastweek':
        # Calculate start (Monday) and end (Sunday) of the previous full week
        current_week_start = today - timedelta(days=today.weekday()) # Monday of current week
        start_date = current_week_start - timedelta(days=7) # Monday of last week
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999) # Sunday of last week
    elif time_period == 'lastmonth':
        # First day of the previous month
        start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # Last microsecond of the previous month
        end_date = today.replace(day=1) - timedelta(microseconds=1)
    elif time_period == 'lastyear':
        start_date = today.replace(year=today.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(year=today.year - 1, month=12, day=31, hour=23, minute=59, second=59, microsecond=999999)
    else: # Default to 'today' if no valid time_period is provided
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    print(f"Calculated Date Range: Start={start_date}, End={end_date}")

    # --- QuerySet for Main Report Data (filtered by all user selections) ---
    # This queryset will populate the main rows of the Excel report
    tickets = Ticket.objects.all()

    # Apply date range filter
    if start_date and end_date:
        tickets = tickets.filter(created_at__range=[start_date, end_date])

    # Apply customer, terminal, and region filters
    if customer_filter != 'all':
        tickets = tickets.filter(terminal__customer__id=customer_filter)
    if terminal_filter != 'all':
        tickets = tickets.filter(terminal__id=terminal_filter)
    if region_filter != 'all':
        tickets = tickets.filter(terminal__region__id=region_filter)

    print(f"Tickets for main report after all filtering: {tickets.count()}")

    # Check if tickets exist before proceeding to avoid empty data
    if tickets.count() == 0:
        print("No tickets found for the main report after filtering.")
        return HttpResponse("No tickets to export matching your criteria.", status=404)

    statistics_base_tickets = Ticket.objects.all()

    # Apply customer, terminal, and region filters to statistics base
    if customer_filter != 'all':
        statistics_base_tickets = statistics_base_tickets.filter(terminal__customer__id=customer_filter)
    if terminal_filter != 'all':
        statistics_base_tickets = statistics_base_tickets.filter(terminal__id=terminal_filter)
    if region_filter != 'all':
        statistics_base_tickets = statistics_base_tickets.filter(terminal__region__id=region_filter)

    # Apply the same date range filter to the statistics base
    if start_date and end_date:
        statistics_base_tickets = statistics_base_tickets.filter(created_at__range=[start_date, end_date])
    
    print(f"Tickets for statistics base after filtering: {statistics_base_tickets.count()}")

    tickets_per_day = [
        statistics_base_tickets.filter(created_at__date=today - timedelta(days=i)).count() 
        for i in range(7)
    ]

    tickets_per_weekday = [
        statistics_base_tickets.filter(created_at__week_day=i).count() 
        for i in range(1, 8)
    ]  

    tickets_per_hour = [
        statistics_base_tickets.filter(created_at__hour=i).count() 
        for i in range(24)
    ]

    tickets_per_month = [
        statistics_base_tickets.filter(created_at__month=i+1).count() 
        for i in range(12)
    ]
    
    
    years = sorted(list(statistics_base_tickets.values_list('created_at__year', flat=True).distinct()))
    tickets_per_year = [
        statistics_base_tickets.filter(created_at__year=year).count() 
        for year in years
    ]
    
    print(f"Tickets per Day (last 7): {tickets_per_day}")
    print(f"Tickets per Weekday: {tickets_per_weekday}")
    print(f"Tickets per Hour: {tickets_per_hour}")
    print(f"Tickets per Month: {tickets_per_month}")
    print(f"Tickets per Year: {tickets_per_year}")


    # Create the Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket Statistics"

    # Add headers to the sheet for main report data
    headers = ['Ticket ID', 'Customer', 'Terminal', 'Region', 'Created At']
    ws.append(headers)

    # Add ticket data rows (from 'tickets' queryset)
    for ticket in tickets:
        customer_name = ticket.terminal.customer.name if ticket.terminal and hasattr(ticket.terminal, 'customer') and ticket.terminal.customer else 'No Customer'
        region_name = ticket.terminal.region.name if ticket.terminal and hasattr(ticket.terminal, 'region') and ticket.terminal.region else 'No Region'
        
       
        created_at_display = ticket.created_at.replace(tzinfo=None) if ticket.created_at else ''

        ws.append([ticket.id, customer_name, ticket.terminal.branch_name, region_name, created_at_display])

    ws.append([]) # Blank row for separation
    ws.append(['--- Statistics Breakdown ---'])

    # Tickets per Day (Last 7 Days)
    ws.append([])
    ws.append(['Tickets per Day (Last 7 Days Relative to Today)'])
    for i in range(7):
        current_date_for_display = today - timedelta(days=i)
        count = tickets_per_day[i]
        ws.append([f"{current_date_for_display.strftime('%Y-%m-%d')}", count])

    # Tickets per Weekday
    ws.append([])
    ws.append(['Tickets per Weekday'])
    
    weekdays_display_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for i in range(len(tickets_per_weekday)): # Loop from 0 to 6
        ws.append([weekdays_display_names[i], tickets_per_weekday[i]])

    # Tickets per Hour
    ws.append([])
    ws.append(['Tickets per Hour'])
    for i, count in enumerate(tickets_per_hour):
        ws.append([f"{i:02d}:00 - {(i+1):02d}:00", count])

    # Tickets per Month
    ws.append([])
    ws.append(['Tickets per Month'])
    months_display_names = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    for i, count in enumerate(tickets_per_month):
        ws.append([months_display_names[i], count])

    # Tickets per Year
    ws.append([])
    ws.append(['Tickets per Year'])
    for i, count in enumerate(tickets_per_year):
        ws.append([years[i], count])


    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ticket_statistics.xlsx"'
    wb.save(response)

    return response

"""
@login_required(login_url='login')
def tickets(request):
    query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    page_number = request.GET.get('page')
    tickets_qs = Ticket.objects.none()

    if request.user.is_authenticated:
        print(f"Authenticated user: {request.user.username}")
        profile = getattr(request.user, 'profile', None)

        # Internal roles (Superusers and staff see all tickets)
        if request.user.is_superuser or request.user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
            print("User has internal access (superuser/staff)")
            tickets_qs = Ticket.objects.all()

        # Overseer role: filter tickets by customer they oversee
        elif Customer.objects.filter(overseer=request.user).exists():
            customer = Customer.objects.filter(overseer=request.user).first()
            print(f"{request.user.username} is Overseer for {customer.name}")
            tickets_qs = Ticket.objects.filter(customer=customer)

        # Custodian role: filter tickets by terminal they are assigned to
        # Custodian role: filter tickets by terminal they are assigned to
        elif profile:
            if profile.terminal:
                print(f"{request.user.username} has terminal: {profile.terminal.branch_name}")
                customer = profile.terminal.customer

                # Ensure the custodian is linked to the correct terminal and customer
                if profile.terminal.custodian == request.user:
                    print(f"{request.user.username} is Custodian for terminal {profile.terminal.branch_name} under customer {customer.name}")
                    # Filter tickets by customer and terminal
                    tickets_qs = Ticket.objects.filter(customer=customer, terminal=profile.terminal)
                else:
                    print(f"{request.user.username} is not the custodian for terminal's customer")
            else:
                print(f"{request.user.username} has a profile but no terminal set")

        else:
            print(f"{request.user.username} has no profile or terminal set")

    # Apply search and status filters if applicable
    if query:
        tickets_qs = tickets_qs.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(problem_category__name__icontains=query)
        )

    if status_filter:
        if status_filter == 'escalated':
            tickets_qs = tickets_qs.filter(is_escalated=True)
        else:
            tickets_qs = tickets_qs.filter(status=status_filter)

    # Order by creation date
    tickets_qs = tickets_qs.order_by('-created_at')

    # Pagination
    paginator = Paginator(tickets_qs, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/tickets.html', {
        'tickets': page_obj,
        'search_query': query,
        'status_filter': status_filter,
    })
"""
@login_required(login_url='login')
def tickets(request):
    # ── 1) Determine user_group & allowed_roles ──
    user_group = None
    allowed_roles = []
    if request.user.groups.exists():
        user_group = request.user.groups.first().name
    if user_group == 'Admin':
        allowed_roles = ['Admin', 'Manager']
    elif user_group == 'Manager':
        allowed_roles = ['Manager', 'Staff']
    else:
        allowed_roles = ['Staff']
    # ── 2) Build your queryset ──
    query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    page_number = request.GET.get('page')
    tickets_qs = Ticket.objects.none()
    if request.user.is_superuser or request.user.groups.filter(name__in=['Director','Manager','Staff']).exists():
        tickets_qs = Ticket.objects.all()
    elif Customer.objects.filter(overseer=request.user).exists():
        customer = Customer.objects.filter(overseer=request.user).first()
        tickets_qs = Ticket.objects.filter(customer=customer)
    else:
        profile = getattr(request.user, 'profile', None)
        if profile and profile.terminal and profile.terminal.custodian == request.user:
            cust = profile.terminal
            tickets_qs = Ticket.objects.filter(customer=cust.customer, terminal=cust)
    # ── 3) Apply filters ──
    if query:
        tickets_qs = tickets_qs.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(problem_category__name__icontains=query)
        )
    if status_filter:
        if status_filter == 'escalated':
            tickets_qs = tickets_qs.filter(is_escalated=True)
        else:
            tickets_qs = tickets_qs.filter(status=status_filter)
    tickets_qs = tickets_qs.order_by('-created_at')
    paginator = Paginator(tickets_qs, 10)
    page_obj = paginator.get_page(page_number)
    # ── 4) Render, now including allowed_roles ──
    return render(request, 'core/helpdesk/tickets.html', {
        'tickets': page_obj,
        'search_query': query,
        'status_filter': status_filter,
        'user_group':    user_group,
        'allowed_roles': allowed_roles,
    })


"""
def create_ticket(request):

    user_group = None  # Default value for user group
    allowed_roles = []  # Default value for allowed_roles

    # Check user groups
    if request.user.groups.exists():
        user_group = request.user.groups.first().name  # Get the first group name of the user

    # Assign roles based on the group or other logic
    if user_group == 'Admin':
        allowed_roles = ['Admin', 'Manager']
    elif user_group == 'Manager':
        allowed_roles = ['Manager', 'Staff']
    else:
        allowed_roles = ['Staff']


    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user

            # Ensure the selected issue is valid and exists
            issue = ticket.title
            if issue and issue not in ISSUE_MAPPING.get(ticket.problem_category.name, []):
                form.add_error('title', 'Selected issue does not belong to the chosen problem category.')

            # Set created_at to custom_created_at if provided
            custom_date = form.cleaned_data.get('custom_created_at')
            if custom_date:
                ticket.created_at = custom_date  # Set custom date

            # Automatically set priority based on the problem category, title, and description
            if not ticket.priority:  # Only set priority if it's not already set
                category = ticket.problem_category
                issue = ticket.title
                description = ticket.description
                ticket.priority = determine_priority(category.name if category else "", issue, description)
                print(f"Priority set to: {ticket.priority}")  # Debugging line
    

            ticket.save()
            return redirect('create_ticket' if 'create_another' in request.POST else 'tickets')
    else:
        terminal_id = request.GET.get('terminal_id')
        form = TicketForm(terminal_id=terminal_id) if terminal_id else TicketForm()

    # Issue mapping for frontend use
    cats = ProblemCategory.objects.all()
    js_mapping = {
        str(cat.pk): ISSUE_MAPPING.get(cat.name, []) for cat in cats
    }

    return render(request, 'core/helpdesk/create_ticket.html', {
        'form': form,
        'issue_mapping': json.dumps(js_mapping),  # Passing the mapping to JS
        'user_group': user_group, # Pass user_group to the template context
        'allowed_roles': allowed_roles  # Add this line
    })
"""
@login_required(login_url='login')
def create_ticket(request):
    # 1) Determine user_group & allowed_roles (same logic as tickets)
    user_group = None
    allowed_roles = []
    if request.user.groups.exists():
        user_group = request.user.groups.first().name
    if user_group == 'Admin':
        allowed_roles = ['Admin', 'Manager']
    elif user_group == 'Manager':
        allowed_roles = ['Manager', 'Staff']
    else:
        allowed_roles = ['Staff']
    # 2) Handle form submission
    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            # Let form.save() compute priority
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            # Optional custom timestamp
            custom_date = form.cleaned_data.get('custom_created_at')
            if custom_date:
                ticket.created_at = custom_date
            ticket.save()
            return redirect('create_ticket' if 'create_another' in request.POST else 'tickets')
    else:
        # Pass terminal_id into the form so __init__ can prefill
        terminal_id = request.GET.get('terminal_id')
        form = TicketForm(terminal_id=terminal_id) if terminal_id else TicketForm()
    # 3) Build JS mapping for dynamic issues dropdown
    cats = ProblemCategory.objects.all()
    js_mapping = { str(cat.pk): ISSUE_MAPPING.get(cat.name, []) for cat in cats }
    return render(request, 'core/helpdesk/create_ticket.html', {
        'form': form,
        'issue_mapping': json.dumps(js_mapping),
        'user_group': user_group,
        'allowed_roles': allowed_roles,
    })


def problem_category(request):
    query = request.GET.get('search', '')
    categories = ProblemCategory.objects.filter(name__icontains=query).order_by('name')
    
    paginator = Paginator(categories, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/helpdesk/problem_category.html', {
        'page_obj': page_obj, 
        'search_query': query,
    })



"""def create_problem_category(request):
    # 1) load all existing categories for the dropdown
    categories = ProblemCategory.objects.order_by('name')
    # 2) see if the user chose one already
    existing_pk = (
        request.GET.get('existing_category') 
        or request.POST.get('existing_category')
    )
    selected_category = None
    if existing_pk:
        selected_category = get_object_or_404(ProblemCategory, pk=existing_pk)
    if request.method == 'POST':
        # 3) bind the form & formset to that instance (or to None for a brand-new)
        form = ProblemCategoryForm(request.POST, instance=selected_category)
        formset = IssueFormSet(request.POST, instance=selected_category)
        if form.is_valid() and formset.is_valid():
            # saves new or updates existing category
            category = form.save()
            # ensure formset knows its FK and save its issues
            formset.instance = category
            formset.save()
            return redirect('problem_category')
    else:
        form = ProblemCategoryForm(instance=selected_category)
        formset = IssueFormSet(instance=selected_category)
    return render(request,
                  'core/helpdesk/create_problem_category.html',
                  {
                      'form': form,
                      'formset': formset,
                      'categories': categories,
                      'selected_category': selected_category,
                  })"""

@user_passes_test(is_director)
def create_problem_category(request):
    if request.method == 'POST':
        print("POST received:", request.POST) 
        form = ProblemCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            print("Category saved!")

            # Redirect based on which button was clicked
            if 'create_another' in request.POST:
                return redirect('create_problem_category')
            return redirect('problem_category')  
        else:
            print("Form errors:", form.errors)
    else:
        form = ProblemCategoryForm()

    return render(request, 'core/helpdesk/create_problem_category.html', {'form': form})




from django.template.loader import render_to_string 
@login_required
def escalate_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Get the current escalation level
    level_order = [level[0] for level in Ticket.ESCALATION_LEVELS]
    
    # Find the current escalation level and the next level
    try:
        current_index = level_order.index(ticket.current_escalation_level) if ticket.current_escalation_level else -1
        next_level = level_order[current_index + 1]  # Next level
    except IndexError:
        messages.warning(request, "Already at the highest escalation level.")
        return redirect('ticket_detail', ticket_id=ticket.id)
    
    if request.method == 'POST':
        form = EscalationNoteForm(request.POST)
        if form.is_valid():
            print("Form is valid")
            note = form.cleaned_data['note']
            print(f"Escalation Note: {note}")  # Debugging line
            
            # Log the escalation to history
            EscalationHistory.objects.create(
                ticket=ticket,
                escalated_by=request.user,
                from_level=ticket.current_escalation_level,
                to_level=next_level,
                note=note
            )

            # Update ticket escalation
            ticket.current_escalation_level = next_level
            ticket.is_escalated = True
            ticket.escalated_at = timezone.now()
            ticket.escalated_by = request.user
            ticket.escalation_reason = note  
            ticket.save()

            # Create a visually appealing HTML email
            subject = f"Ticket #{ticket.id} Escalated to {ticket.current_escalation_level}"
            html_message = render_to_string('core/helpdesk/ticket_escalated.html', {
                'ticket': ticket,
                'note': note,
                'next_level': next_level
            })
            
            send_mail(
                subject=subject,
                message="This is an HTML email. Please view it in an HTML-compatible email client.",
                from_email="godblessodhiambo@gmail.com",
                recipient_list=get_email_for_level(next_level),
                fail_silently=False,
                html_message=html_message
            )

            messages.success(request, f"Ticket has been escalated to {next_level}.")
            return redirect('ticket_detail', ticket_id=ticket.id)
    else:
        print("There is a problem")
        form = EscalationNoteForm()

    return render(request, 'core/helpdesk/escalate_ticket.html', {
        'ticket': ticket,
        'form': form,
        'next_level': next_level
    })

def get_email_for_level(level):
    # This function fetches emails based on escalation level from settings.
    return settings.ESCALATION_LEVEL_EMAILS.get(level, [])

def notify_group(level, ticket):
    # Example: Notify the next escalation level via email
    email_recipient = get_email_for_level(level)  # Define how to get the right email for the group
    
    send_mail(
        f'Ticket #{ticket.id} has been escalated to {level}',
        f'The ticket with the issue "{ticket.title}" has been escalated to {level}.',
        settings.DEFAULT_FROM_EMAIL,
        [email_recipient],
        fail_silently=False
    )

@login_required(login_url='login')
def get_terminal_details(request, terminal_id):
    try:
        terminal = Terminal.objects.get(id=terminal_id)
        response_data = {
            'customer_id': terminal.customer.id if terminal.customer else None,
            'region_id': terminal.region.id if terminal.region else None,
        }
        return JsonResponse(response_data)
    except Terminal.DoesNotExist:
        return JsonResponse({'error': 'Terminal not found'}, status=404)

def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    comments = ticket.comments.order_by('-created_at')
    form = TicketEditForm(instance=ticket)
    comment_form = TicketCommentForm()

    is_manager = request.user.groups.filter(name='Manager').exists()

    # Fetch staff members for the dropdown (only for managers)
    staff_users = User.objects.filter(groups__name='Staff')

    if request.method == 'POST':
        if 'add_comment' in request.POST:
            comment_form = TicketCommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.ticket = ticket
                comment.created_by = request.user
                comment.save()
                return redirect('ticket_detail', ticket_id=ticket.id)

        elif 'edit_ticket' in request.POST:
            form = TicketEditForm(request.POST, instance=ticket)
            if form.is_valid():
                form.save()
                return redirect('ticket_detail', ticket_id=ticket.id)

        elif 'assign_ticket' in request.POST and is_manager:
            staff_id = request.POST.get('assigned_to')
            if staff_id:
                staff_member = get_object_or_404(User, id=staff_id, groups__name='Staff')
                ticket.assigned_to = staff_member
                ticket.save()

                subject = f"🎫 Ticket #{ticket.id} Assigned to You"
                text_content = (
                    f"Hello {staff_member.get_full_name() or staff_member.username},\n\n"
                    f"You have been assigned ticket #{ticket.id} - {ticket.title}.\n"
                    f"Please log in to the system to view and resolve it:\n"
                    f"{request.build_absolute_uri(reverse('ticket_detail', args=[ticket.id]))}\n\n"
                    "Thank you."
                )

                html_content = f"""
                <html>
                <body style="font-family: Arial, sans-serif; background-color: #f4f6f8; margin:0; padding:0;">
                <table width="100%" cellpadding="0" cellspacing="0" 
                    style="max-width:600px; margin:40px auto; background:#ffffff; border-radius:8px;
                            overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,0.08);">
                    
                    <!-- LOGO HEADER -->
                    <tr>
                        <td style="background-color:#ffffff; text-align:center; padding:20px;">
                            <img src="cid:logo" alt="BRITS Logo" width="150" 
                                style="display:block; margin:0 auto;" />
                        </td>
                    </tr>

                    <!-- TITLE BAR -->
                    <tr>
                        <td style="background-color:#172d69; text-align:center; padding:15px;">
                            <h1 style="color:#ffffff; font-size:24px; margin:0;">
                                📌 New Ticket Assigned
                            </h1>
                        </td>
                    </tr>

                    <!-- BODY -->
                    <tr>
                        <td style="padding:30px;">
                            <p style="font-size:16px; line-height:1.6; margin-bottom:20px;">
                                Hello <strong>{staff_member.get_full_name() or staff_member.username}</strong>,
                            </p>
                            <p style="font-size:16px; line-height:1.6; margin-bottom:20px;">
                                You have been assigned a new ticket with the following details:
                            </p>

                            <!-- TICKET DETAILS TABLE -->
                            <table cellpadding="0" cellspacing="0" width="100%" 
                                style="margin-bottom:30px; border:1px solid #e0e0e0; border-radius:6px; overflow:hidden;">
                                <tr>
                                    <td style="padding:10px; background-color:#f9fafc; border-bottom:1px solid #e0e0e0;">
                                        <strong>ID:</strong>
                                    </td>
                                    <td style="padding:10px; border-bottom:1px solid #e0e0e0;">{ticket.id}</td>
                                </tr>
                                <tr>
                                    <td style="padding:10px; background-color:#f9fafc; border-bottom:1px solid #e0e0e0;">
                                        <strong>Title:</strong>
                                    </td>
                                    <td style="padding:10px; border-bottom:1px solid #e0e0e0;">{ticket.title}</td>
                                </tr>
                                <tr>
                                    <td style="padding:10px; background-color:#f9fafc;">
                                        <strong>Status:</strong>
                                    </td>
                                    <td style="padding:10px;">{ticket.status}</td>
                                </tr>
                            </table>

                            <!-- CTA BUTTON -->
                            <p style="text-align:center; margin-bottom:30px;">
                                <a href="{request.build_absolute_uri(reverse('ticket_detail', args=[ticket.id]))}"
                                style="background-color:#007bff; color:#ffffff; text-decoration:none;
                                        padding:14px 24px; border-radius:6px; display:inline-block;
                                        font-size:16px; font-weight:bold; box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                                    View & Resolve Ticket
                                </a>
                            </p>
                        </td>
                    </tr>

                    <!-- FOOTER -->
                    <tr>
                        <td style="background-color:#f0f2f5; padding:15px; text-align:center; font-size:12px; color:#7f8c8d;">
                            This is an automated message. Please do not reply.<br>
                            <span style="color:red;">⚠ Urgent tickets must be addressed promptly.</span>
                        </td>
                    </tr>
                </table>
                </body>
                </html>
                """

                msg = EmailMultiAlternatives(
                    subject,
                    text_content,
                    settings.DEFAULT_FROM_EMAIL,
                    [staff_member.email]
                )
                msg.mixed_subtype = 'related'
                msg.attach_alternative(html_content, "text/html")

                logo_path = os.path.join(settings.BASE_DIR, 'static', 'icons', 'logo.png')
                with open(logo_path, 'rb') as f:
                    logo = MIMEImage(f.read())
                    logo.add_header('Content-ID', '<logo>')
                    logo.add_header('Content-Disposition', 'inline; filename="logo.png"')
                    msg.attach(logo)

                msg.send()
                return redirect('ticket_detail', ticket_id=ticket.id)


    context = {
        'ticket': ticket,
        'form': form,
        'comments': comments,
        'comment_form': comment_form,
        'is_admin': request.user.is_superuser,
        'is_editor': request.user.groups.filter(name='Editor').exists(),
        'can_resolve': request.user.groups.filter(name='Resolver').exists(),
        'is_manager': is_manager,
        'staff_users': staff_users if is_manager else None
    }

    return render(request, 'core/helpdesk/ticket_detail.html', context)

@login_required
def edit_comment(request, comment_id):
    comment = get_object_or_404(TicketComment, id=comment_id)

    if request.user != comment.created_by and not request.user.is_superuser:
        messages.error(request, "You don't have permission to edit this comment.")
        return redirect('ticket_detail', ticket_id=comment.ticket.id)

    if request.method == 'POST':
        form = TicketCommentForm(request.POST, instance=comment)
        if form.is_valid():
            form.save()
            messages.success(request, "Comment updated successfully.")
            return redirect('ticket_detail', ticket_id=comment.ticket.id)
    else:
        form = TicketCommentForm(instance=comment)

    return render(request, 'core/helpdesk/edit_comment.html', {'form': form, 'comment': comment})


@login_required
def delete_comment(request, comment_id):
    comment = get_object_or_404(TicketComment, id=comment_id)

    if request.user != comment.created_by and not request.user.is_superuser:
        messages.error(request, "You don't have permission to delete this comment.")
        return redirect('ticket_detail', ticket_id=comment.ticket.id)

    if request.method == 'POST':
        ticket_id = comment.ticket.id
        comment.delete()
        messages.success(request, "Comment deleted.")
        return redirect('ticket_detail', ticket_id=ticket_id)



@login_required
def resolve_ticket_view(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    resolution = request.POST.get('resolution', '').strip()

    # Check if the user is authorized to resolve the ticket
    if is_director(request.user) or is_manager(request.user) or is_staff(request.user):
        if ticket.status != 'resolved':
            ticket.resolution = resolution
            ticket.status = 'closed'
            ticket.resolved_by = request.user  
            ticket.resolved_at = timezone.now()
            ticket.save()
            messages.success(request, 'Ticket resolved successfully!')
            return redirect('ticket_detail', ticket_id=ticket.id)
        else:
            messages.error(request, 'Ticket already resolved')
            return render(request, 'core/helpdesk/error.html')

    elif request.user.has_perm('can_resolve_ticket'):
        # Custom permission check
        if ticket.status != 'resolved':
            ticket.status = 'resolved'
            ticket.resolved_by = request.user  
            ticket.resolved_at = timezone.now() 
            ticket.save()
            messages.success(request, 'Ticket resolved successfully!')
            return redirect('ticket_detail', ticket_id=ticket.id)
        else:
            messages.error(request, 'Ticket already resolved!')
            return render(request, 'core/helpdesk/error.html')

    # If the user doesn't have permission
    messages.error(request, 'You do not have permission to resolve this ticket.')
    return render(request, 'core/helpdesk/permission_denied.html')


@user_passes_test(is_director)
def delete_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    ticket.delete()
    messages.success(request, "Ticket deleted successfully.")
    return redirect('tickets')

def ticket_statuses(request):
    user = request.user
    is_custodian = user.groups.filter(name="Custodian").exists()
    is_overseer = user.groups.filter(name="Overseer").exists()
    is_customer = user.groups.filter(name="Customer").exists()

    return render(request, 'core/helpdesk/ticket_statuses.html', {
        "is_custodian": is_custodian,
        "is_overseer": is_overseer,
        "is_customer": is_customer,
    })


@login_required
def tickets_by_status(request, status):
    tickets_qs = Ticket.objects.filter(
        status__iexact=status.replace('-', '_')
    ).select_related('customer', 'terminal')

    user = request.user
    user_profile = getattr(user, 'profile', None) 

    # 1. Internal roles (Superusers and specific staff groups see all tickets)
    if user.is_superuser or user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
        print("User has internal access (superuser/staff) - viewing all tickets for this status.")

    elif Customer.objects.filter(overseer=user).exists():
        customer_overseen = Customer.objects.filter(overseer=user).first()
        if customer_overseen:
            print(f"{user.username} is Overseer for customer: {customer_overseen.name}")
            tickets_qs = tickets_qs.filter(customer=customer_overseen)
        else:
            tickets_qs = Ticket.objects.none()
            print(f"{user.username} is Overseer but no customer found (unexpected).")

    elif user_profile and user_profile.terminal:
        if user_profile.terminal.custodian == user:
            print(f"{user.username} is Custodian for terminal: {user_profile.terminal.branch_name}")
            tickets_qs = tickets_qs.filter(terminal=user_profile.terminal)
        else:
            tickets_qs = Ticket.objects.none()
            print(f"{user.username} has terminal in profile but is not its custodian. Returning no tickets.")
    else:
        tickets_qs = Ticket.objects.none()
        print(f"User {user.username} does not match any specific access role. Returning no tickets.")

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(tickets_qs, 10) 

    try:
        tickets = paginator.page(page)
    except PageNotAnInteger:
        tickets = paginator.page(1)
    except EmptyPage:
        tickets = paginator.page(paginator.num_pages)

    print(f"Final tickets count for {status} status: {tickets_qs.count()}")  

    return render(request, 'core/helpdesk/ticket_by_status.html', {
        'status': status.title().replace('-', ' '),
        'tickets': tickets,
        'paginator': paginator
    })


@user_passes_test(is_director)
def edit_problem_category(request, category_id):
    category = get_object_or_404(ProblemCategory, pk=category_id)
    if request.method == 'POST':
        form = ProblemCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('problem_category')
    else:
        form = ProblemCategoryForm(instance=category)

    return render(request, 'core/helpdesk/edit_problem_category.html', {'form': form})

@user_passes_test(is_director)
def delete_problem_category(request, category_id):
    category = get_object_or_404(ProblemCategory, id=category_id)
    category.delete()
    messages.success(request, "Problem category deleted successfully.")
    return redirect('problem_category')

# Master Data Views
def customers(request):
    if request.method == "POST" and request.FILES.get("file"):
        csv_file = request.FILES["file"]
        decoded_file = csv_file.read().decode("utf-8").splitlines()
        reader = csv.DictReader(decoded_file)

        for row in reader:
            name = row.get("name", "").strip()
            if name: 
                Customer.objects.create(name=name)

        messages.success(request, "Customers uploaded successfully!")

    # Pagination setup
    all_customers = Customer.objects.exclude(name__exact="").exclude(name__isnull=True).order_by('id')
    paginator = Paginator(all_customers, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "core/helpdesk/customers.html", {"customers": page_obj})

@user_passes_test(is_director)
def create_customer(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            Customer.objects.create(name=name)
            messages.success(request, "Customer added successfully.")
            return redirect("customers")
        else:
            messages.error(request, "Customer name is required.")

    return render(request, "core/helpdesk/create_customer.html")

@user_passes_test(is_director)
def delete_customer(request, id):
    customer = get_object_or_404(Customer, id=id)
    customer.delete()
    messages.success(request, "Customer deleted successfully.")
    return redirect('customers')

@login_required(login_url='login')
def regions(request):
    if request.method == 'POST':
        name = request.POST.get('region_name')
        if name:
            Region.objects.create(name=name)
            return redirect('regions')

    # Fetch all regions
    all_regions = Region.objects.all().order_by('id')

    # Pagination setup
    paginator = Paginator(all_regions, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/regions.html', {'regions': page_obj})

@user_passes_test(is_director)
def delete_region(request, region_id):
    region = get_object_or_404(Region, id=region_id)
    region.delete()
    messages.success(request, "Region deleted successfully.")
    return redirect('regions')

@login_required(login_url='login')
def terminals(request):
    form = TerminalForm()
    upload_form = TerminalUploadForm()

    # Get all the required objects to pass to the template
    customers = Customer.objects.all()
    regions = Region.objects.all()
    zones = Zone.objects.all()

    # Handle POST request for terminal creation or CSV upload
    if request.method == 'POST':
        # Handle terminal creation
        if 'create' in request.POST or 'create_another' in request.POST:
            print("Form submitted")
            form = TerminalForm(request.POST)
            if form.is_valid():
                print("Form is valid")
                try:
                    form.save()
                    messages.success(request, "Terminal created successfully.")

                    # Handle the "Create & Add Another" functionality
                    if 'create_another' in request.POST:
                        return redirect('terminals')

                    return redirect('terminals')  
                except Exception as e:
                    messages.error(request, f"Error creating terminal: {e}")
                    print(f"Error creating terminal: {e}")
            else:
                print("Form is not valid")
                print("Form errors:", form.errors) 
        # Handle CSV upload for terminals
        elif 'upload_file' in request.POST:
            upload_form = TerminalUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                file = upload_form.cleaned_data['file']
                try:
                    if file.name.endswith('.csv'):
                        df = pd.read_csv(file)
                    else:
                        df = pd.read_excel(file)

                    # Process each row and create terminal objects
                    for _, row in df.iterrows():
                        Terminal.objects.create(
                            customer=Customer.objects.get(name=row['customer']),
                            branch_name=row['branch_name'],
                            cdm_name=row['cdm_name'],
                            serial_number=row['serial_number'],
                            region=Region.objects.get(name=row['region']),
                            model=row['model'],
                            zone=Zone.objects.get(name=row['zone']),
                        )
                    messages.success(request, "Terminals imported successfully.")
                except Exception as e:
                    messages.error(request, f"Error importing file: {e}")
                return redirect('terminals')

    # GET request: Display the page with all terminals
    query = request.GET.get('q', '').strip()
    all_terminals = Terminal.objects.all().order_by('id') 
    if query:
        all_terminals = all_terminals.filter(
            Q(customer__name__icontains=query) |
            Q(branch_name__icontains=query) |
            Q(cdm_name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(region__name__icontains=query) |
            Q(model__icontains=query) |
            Q(zone__name__icontains=query)
        )
    paginator = Paginator(all_terminals, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the required objects to the template
    return render(request, 'core/helpdesk/terminals.html', {
        'form': form,
        'upload_form': upload_form,
        'terminals': page_obj,
        'customers': customers,
        'regions': regions,
        'zones': zones
    })


@require_POST
def edit_terminal(request, terminal_id):
    terminal = get_object_or_404(Terminal, id=terminal_id)
    terminal.customer_id = request.POST.get('customer')
    terminal.branch_name = request.POST.get('branch_name')
    terminal.cdm_name = request.POST.get('cdm_name')
    terminal.serial_number = request.POST.get('serial_number')
    terminal.region_id = request.POST.get('region')
    terminal.model = request.POST.get('model')
    terminal.zone_id = request.POST.get('zone')
    
    try:
        terminal.save()
        messages.success(request, "Terminal updated successfully.")
    except Exception as e:
        messages.error(request, f"Error updating terminal: {e}")
    
    return redirect('terminals')

@login_required(login_url='login')
def fetch_tickets(request, terminal_id):
    tickets = Ticket.objects.filter(terminal_id=terminal_id).values('id', 'title')
    
    if tickets:
        return JsonResponse({"success": True, "tickets": list(tickets)})
    else:
        return JsonResponse({"success": False, "message": "No tickets found."})


@user_passes_test(is_director)
def delete_terminal(request, terminal_id):
    terminal = get_object_or_404(Terminal, id=terminal_id)
    terminal.delete()
    messages.success(request, "Terminal removed successfully.")
    return redirect('terminals')

def units(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        if name and description:
            Unit.objects.create(name=name, description=description)
        return redirect('units')

    all_units = Unit.objects.all().order_by('id')
    
    paginator = Paginator(all_units, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/units.html', {'page_obj': page_obj})

@user_passes_test(is_director)
def delete_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    unit.delete()
    messages.success(request, "Unit removed successfully.")
    return redirect('units')

@login_required(login_url='login')
def system_users(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        role = request.POST.get('role')
        users = User.objects.all()
        if username and email and role:
            SystemUser.objects.create(username=username, email=email, role=role)
        return redirect('system_users')

    all_users = User.objects.all().order_by('id')
    # Add pagination: Show 10 users per page
    paginator = Paginator(all_users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/helpdesk/users.html', {'page_obj': page_obj})

@user_passes_test(is_director)
def delete_system_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.user == user:
        messages.error(request, "You cannot delete your own account.")
    else:
        user.delete()
        messages.success(request, "User deleted successfully.")
    return redirect('system_users')

@login_required(login_url='login')
def zones(request):
    if request.method == 'POST':
        name = request.POST.get('name')

        if name: 
            Zone.objects.create(name=name)
            messages.success(request, "Zone created successfully.")
            return redirect('zones')
        else:
            messages.error(request, "Name is required.")

    all_zones = Zone.objects.all().order_by('id')

    # Add pagination: Show 10 zones per page
    paginator = Paginator(all_zones, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/zones.html', {
        'page_obj': page_obj,
    })

@user_passes_test(is_director)
def delete_zone(request, zone_id):
    zone = get_object_or_404(Zone, id=zone_id)
    zone.delete()
    messages.success(request, "Zone deleted successfully.")
    return redirect('zones') 

@login_required(login_url='login')
def reports(request):
    user_group = None
    customers = Customer.objects.all()
    terminals = Terminal.objects.all()

    # Correct way to get a single terminal for the custodian
    terminals = Terminal.objects.filter(custodian=request.user)  # Still returns a QuerySet
    if terminals.exists():
        terminal = terminals.first() 
        user_group = "Custodian"
        terminal = Terminal.objects.filter(custodian=request.user)
        customers = Customer.objects.filter(
            id__in=terminal.values_list('customer_id', flat=True)
        ) # Get the first terminal if it exists
    # Custodian logic
    #if Terminal.objects.filter(custodian=request.user).exists():
        
    elif Customer.objects.filter(overseer=request.user).exists():
        user_group = "Overseer"
        assigned_customers = Customer.objects.filter(overseer=request.user)
        customers = assigned_customers  # Only customers assigned to the overseer

    # Base ticket query
    tickets = Ticket.objects.prefetch_related('comments').all().order_by('-created_at')

    # Apply user‐level filtering
    if user_group == "Custodian":
        tickets = tickets.filter(terminal__in=terminals)
    elif user_group == "Overseer":
        tickets = tickets.filter(customer__in=customers)

    customer = request.GET.get('customer')
    terminal_name = request.GET.get("terminal_name")
    region = request.GET.get('region')
    category = request.GET.get('category')

    filter_by_customer = False
    filter_by_terminal = False

    if customer and customer != 'All' and customer != "None":
        tickets = tickets.filter(customer_id=customer)
        filter_by_customer = True

    if terminal_name:
        tickets = tickets.filter(terminal__branch_name__icontains=terminal_name)
        filter_by_terminal = True

    if region and region != 'All' and region != "None":
        tickets = tickets.filter(region_id=region)

    if category and category != 'All' and category != "None":
        tickets = tickets.filter(problem_category_id=category)

    # Date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        tickets = tickets.filter(created_at__date__gte=parse_date(start_date))
    if end_date:
        tickets = tickets.filter(created_at__date__lte=parse_date(end_date))

    # Export to Excel logic
    if request.GET.get('download') == 'excel':
        customer_name = Customer.objects.get(id=customer).name if customer and customer not in ['All', 'None'] else None
        terminal_filter = terminal_name if terminal_name else None

        return export_tickets_to_excel(
            tickets,
            include_terminal=filter_by_customer,
            customer_name=customer_name,
            terminal_name=terminal_filter,
            start_date=start_date,
            end_date=end_date
        )

    # Pagination
    paginator = Paginator(tickets, 10)
    page = request.GET.get('page')

    try:
        tickets_page = paginator.page(page)
    except PageNotAnInteger:
        tickets_page = paginator.page(1)
    except EmptyPage:
        tickets_page = paginator.page(paginator.num_pages)

    # Context to be passed to the template
    context = {
        'tickets': tickets_page,
        'page_obj': tickets_page,
        'customers': customers,
        'terminals': terminals,
        'regions': Region.objects.all(),
        'categories': ProblemCategory.objects.all(),
        'filter_by_customer': filter_by_customer,
        'filter_by_terminal': filter_by_terminal,
        'user_group': user_group,  # Pass user group for conditionally disabling filters
    }

    return render(request, 'core/helpdesk/reports.html', context)

def export_tickets_to_excel(tickets, include_terminal=False, customer_name=None, terminal_name=None, start_date=None, end_date=None):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from django.http import HttpResponse

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Tickets'

    headers = []
    if include_terminal:
        headers.append('Terminal')

    headers += ['Created At', 'Updated At', 'Problem Category', 'Status', 'Responsible', 'Description', 'Comments', 'Resolution']
    sheet.append(headers)

    for ticket in tickets:
        comments_text = "\n".join([
            f"{comment.created_by.username if comment.created_by else 'Unknown'} ({comment.created_at.strftime('%Y-%m-%d')}): {comment.content}"
            for comment in ticket.comments.all()
        ])
        row = []
        if include_terminal:
            row.append(ticket.terminal.branch_name)  
        row += [
            ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            ticket.updated_at.strftime('%Y-%m-%d %H:%M'),
            str(ticket.problem_category),
            ticket.status,
            str(ticket.responsible),
            ticket.description,
            comments_text,
            ticket.resolution or "",
        ]
        sheet.append(row)

    comment_col_index = len(headers)
    for row in sheet.iter_rows(min_row=2, min_col=comment_col_index, max_col=comment_col_index):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    name_part = "report"
    if customer_name:
        name_part = f"{customer_name.replace(' ', '_')}_report"
    elif terminal_name:
        name_part = f"{terminal_name.replace(' ', '_')}_report"

    date_part = ''
    if start_date and end_date:
        date_part = f"{start_date}_to_{end_date}"
    elif start_date:
        date_part = f"from_{start_date}"
    elif end_date:
        date_part = f"to_{end_date}"

    filename = f"{name_part}_{date_part or timezone.now().strftime('%Y-%m-%d')}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)

    return response


@login_required(login_url='login')
def version_controls(request):
    print("view reached") 
    form = VersionControlForm()

    if request.method == 'POST':
        if 'create' in request.POST or 'create_another' in request.POST:
            form = VersionControlForm(request.POST)
            if form.is_valid():
                print("form is valid")
                form.save()
                if 'create_another' in request.POST:
                    form = VersionControlForm()
                else:
                    return redirect('version_controls')
            else:
                print("Form is not valid")
    versions = VersionControl.objects.all().order_by('-created_at')

    # Handle AJAX filtering
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        terminal = request.GET.get('terminal')
        firmware = request.GET.get('firmware')
        app_version = request.GET.get('app_version')
        manufacturer = request.GET.get('manufacturer')

        if terminal and terminal != 'All':
            versions = versions.filter(terminal__id=terminal)
        if manufacturer and manufacturer != 'All':
            versions = versions.filter(manufacturer=manufacturer)
        

        return render(request, 'core/helpdesk/partials/version_table.html', {
            'versions': versions
        })

    # Paginate the full list
    paginator = Paginator(versions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Terminal filter options
    terminals = VersionControl.objects.select_related('terminal').values(
        'terminal__branch_name', 'terminal__id'
    ).distinct()

    context = {
        'form': form,
        'page_obj': page_obj,
        'versions': page_obj,  # This line ensures compatibility with both paginated and non-paginated loops
        'terminals': terminals,
        'manufacturers': VersionControl.objects.values_list('manufacturer', flat=True).distinct(),
        #'firmwares': VersionControl.objects.values_list('firmware', flat=True).distinct(),
        #'app_versions': VersionControl.objects.values_list('app_version', flat=True).distinct(),
    }
    return render(request, 'core/helpdesk/version_control.html', context)

@login_required(login_url='login')
def version_detail(request, pk):
    version = get_object_or_404(VersionControl, pk=pk)
    comments = version.comments.all().order_by('-created')  # Latest first

    if request.method == 'POST':
        comment_text = request.POST.get('comment')
        if comment_text:
            VersionComment.objects.create(version=version, text=comment_text)
        return redirect('version_detail', pk=pk)

    return render(request, 'core/helpdesk/version_detail.html', {
        'version': version,
        'comments': comments,
        
    })

@login_required(login_url='login')
def edit_version(request, pk):
    version = get_object_or_404(VersionControl, pk=pk)
    if request.method == 'POST':
        form = VersionControlForm(request.POST, instance=version)
        if form.is_valid():
            form.save()
            return redirect('version_detail', pk=pk)
    else:
        form = VersionControlForm(instance=version)

    return render(request, 'core/helpdesk/edit_version.html', {'form': form, 'version': version})

@user_passes_test(is_director)
def delete_version(request, pk):
    version = get_object_or_404(VersionControl, pk=pk)
    version.delete()
    return redirect('version_controls') 